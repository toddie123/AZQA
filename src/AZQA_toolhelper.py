"""""
Aztech Locknut Company Tool Tracker Program
This program is licensed to Aztech Engineering Inc. and its subsidiaries

Written by Todd Kuebelbeck, 2021 developed at and for Aztech Locknut Company
--------------------------------------------------------------------------------
Helper class for Aztech Tool Tracker; Tool class
"""""

from openpyxl import load_workbook  # Importing Open Python Excel library for spreadsheet manipulation




class Tool(object):

    """"
    Create tool instance variables
    """
    def __init__(self, ID, spreadsheet_path):

        self.ID = ID
        self.spreadsheet = spreadsheet_path
        self.tool_type = ""
        self.use_limit = ""
        self.cal_date = ""
        self.cal_exp = ""
        self.log_path = ""
        self.cert_path = ""
        self.sheetrow = ""  # row that contains data for ID of interest

        wb = load_workbook(filename=spreadsheet_path)   # create workbook object from passedthrough filepath
        ws = wb.active  # create worksheet object that is the active sheet of the excel book

        self.ID_col = "C"    # Col that contains IDs default from sheet as of 02/01/2021 can be set externally

    def set_spreadsheet_path(self, spread_path_intput):
        self.spreadsheet = spread_path_intput

    def find_row(self):

        for row in self.ws.iter_rows(self.ID_col):
            for cell in row:
                if cell.value == self.ID:
                    self.sheetrow = cell.row

                # TODO ADD TERMINATOR

    def set_tool_info(self):

        if self.wb.cell(row = self.sheetrow, column = self.ID_col).value is self.ID:
            # Creating the name of the tool. Concatenates description (col E) and model (col G)
            # TODO create column variables for parameters
            self.tool_type = self.wb.cell(row = self.sheetrow, column = 'E').value + " " + self.wb.cell(row = self.sheetrow, column = 'G')

            self.use_limit = self.wb.cell(row = self.sheetrow, column = 'P').value
            self.cal_date = self.wb.cell(row = self.sheetrow, column = 'M').value
            self.cal_exp = self.wb.cell(row = self.sheetrow, column = 'L').value
            self.log_path = self.wb.cell(row = self.sheetrow, column = 'Q').hyperlink.target
            self.cert_path = self.wb.cell(row = self.sheetrow, column = 'B').hyperlink.target










